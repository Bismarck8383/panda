import openpyxl
import mysql.connector
from configuracion import DATABASE_CONFIG

def insert_data_to_database(data, table_name):
    cnx = mysql.connector.connect(**DATABASE_CONFIG)
    cursor = cnx.cursor()
    inserted_rows = 0

    for row in data:
        # Reemplaza los valores None por una cadena vacía y completa las filas faltantes
        row = [value if value is not None else '' for value in row] + [''] * (17 - len(row))
        row = row[:17]  # Asegúrate de que la lista tenga exactamente 17 elementos

        sql = f"""INSERT INTO `{table_name}` 
                (Nombre_Fiscal, Nombre_Comercial, Pais, CIF_NIF, Tipo, Cuenta_Contable, Contacto, Cliente_Principal, Telefono, Direccion, Poblacion, Provincia, Cod_Postal, Email, Enviar_publicidad, Locales, Grupos) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        cursor.execute(sql, tuple(row))
        cnx.commit()
        inserted_rows += 1

    cursor.close()
    cnx.close()
    return inserted_rows

# Leer datos del archivo xlsx
excel_data = openpyxl.load_workbook("archivos/convertido.xlsx")
dataframe = excel_data[excel_data.sheetnames[0]]  # Selecciona la primera hoja

data = []
for row in range(2, dataframe.max_row + 1):
    _row = []

    for col in range(1, dataframe.max_column + 1):
        _row.append(dataframe.cell(row=row, column=col).value)

    data.append(_row)

# Insertar datos en la base de datos
table_name = "clientes"
inserted_rows = insert_data_to_database(data, table_name)

# Mensaje de confirmación
print(f"Se han insertado {inserted_rows} filas en la tabla '{table_name}' correctamente.")
