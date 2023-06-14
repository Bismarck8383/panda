import datetime
import openpyxl
import mysql.connector
from configuracion import DATABASE_CONFIG

def convertir_a_fecha(fecha_str):
    if isinstance(fecha_str, int):
        return str(fecha_str)  # Devuelve el valor entero como cadena
    else:
        fecha_str = str(fecha_str)  # Convierte el valor a cadena
        return fecha_str

def insertar_datos_en_base_de_datos(datos, nombre_tabla):
    cnx = mysql.connector.connect(**DATABASE_CONFIG)
    cursor = cnx.cursor()
    filas_insertadas = 0

    for fila in datos:
        # Reemplaza los valores None por una cadena vacía y completa las filas faltantes
        fila = [valor if valor is not None else '' for valor in fila] + [''] * (22 - len(fila))
        fila = fila[:22]  # Asegúrate de que la lista tenga exactamente 22 elementos

        # Convierte la cadena de fecha al formato adecuado para la base de datos
        fecha = convertir_a_fecha(fila[0])

        sql = f"""INSERT INTO `{nombre_tabla}`
                (Fecha, Numero, Local, TPV, Centro, Ubicacion, Usuario, Base, Cuota, Total, Propina, FechaCreacion,
                TipoDoc, FormaPago, Cliente, Pais, CIF_NIF, CuentaContable, Direccion, Poblacion, Provincia, CodigoPostal)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        cursor.execute(sql, (convertir_a_fecha(fecha),) + tuple(fila[1:]))
        cnx.commit()
        filas_insertadas += 1

    cursor.close()
    cnx.close()
    return filas_insertadas

# Leer datos del archivo XLSX
excel_data = openpyxl.load_workbook("archivos/convertido2805.xlsx")
hoja = excel_data[excel_data.sheetnames[0]]  # Selecciona la primera hoja

datos = []
for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, values_only=True):
    datos.append(fila)

# Insertar datos en la base de datos
nombre_tabla = "factura"
filas_insertadas = insertar_datos_en_base_de_datos(datos, nombre_tabla)

# Mensaje de confirmación
print(f"Se han insertado {filas_insertadas} filas en la tabla '{nombre_tabla}' correctamente.")

#coment all